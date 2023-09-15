from django.shortcuts import render
import requests
import json
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
# from webdriver_manager.core.utils import ChromeType
options = Options()
options.add_experimental_option("detach", True)


options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--disable-extensions")
options.add_argument("--disable-popup-blocking")


options.add_argument("--disable-infobars")  # Disable the "Chrome is being controlled by automated test software" message
options.add_argument("--start-maximized")  # Start the browser in maximized mode
options.add_argument("--no-sandbox")  # Disable the sandbox mode
options.add_argument("--disable-dev-shm-usage")  # Disable the /dev/shm usage
options.add_argument("--disable-gpu") 


# options.add_argument("--proxy-server=167.114.170.75:56084")

import socket

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait


from django.utils.datastructures import MultiValueDictKeyError
from pyexcel_xls import get_data as xls_get
from pyexcel_xlsx import get_data as xlsx_get
import pandas as pd
import re
import time
import random

from bs4 import BeautifulSoup

import os
# webdriver.DesiredCapabilities.CHROME['acceptSslCerts']=True
# from getproxies import get_proxies
# proxies = get_proxies()
# PROXY = random.choice(proxies)
import csv
import requests
import openpyxl
# wb.save(filename='Test.xlsx')

# file_path = "D:\\job source project\\JobManagement\\Skills.txt"
# with open(file_path, "r") as file:
#     SKILLS = file.read().splitlines()
    
# print(SKILLS)
# SKILLS = ['Java', 'Python', 'Software Engineer', 'Project Manager', 'Product Manager', 'Software Developer',
#           'Software Engineering', 'Full Stack', 'Quality Engineer', 'AWS', 'Solution Architect', 'Solutions Architect',
#           'Customer Service Respresentative', 'Analyst', 'Springboot', 'Auditor', '.Net', 'Oracle', 'Database', 'DBA',
#           'DevOps', 'Technical Lead', 'Salesforce', 'Developer', 'ServiceNow', 'Big Data', 'Application Developer',
#           'Technical Architect', 'Cloud','Drupal', 'Software Tester', 'C#', 'Sales Engineer', 'Project Engineer']
# ['montana.edu', 'stripe.com', '1800flowers.com', 'bitsgap.com', 'pointclickcare.com', 'brothers-brick.com', 
#  'sevenrooms.com', '2ndswing.com', '2ndwatch.com', '3boxlabs.com', '3dsystems.com', 'coresystems.hu', '3i-infotech.com',
#  '4cgeoworks.com', 'patriotsoftware.com', '32degrees.com', 'kingautomotive.net', 'elevenwarriors.com', 'facebook.com',
#  'summitsalon.com', '17lumberandrental.com', 'xhamster.com', 'adobe.com', 'apvit.com', 'aplacetocallhomeilp.org', 
#  'aws-startup-lofts.com', 'copierserviceswestcovina.com', 'aaa.com', 'aaafoundationservice.com', 'aaratechnologies.com', 
#  'aaratechnologies.com', 'aandbcleaningservices.co.uk', 'abc7.com', 'abctechnologies.com', 'abtassociates.com', 'ace-limited.info', 
#  'aciinfotech.com', 'acvauctions.com', 'adbsafegate.com', 'ad.net', 'advantech.com', 'ae.com', 'aegworldwide.com', 'aesfluids.com',
#  'geodetics.com', 'ags.gob.mx', 'aheadofthyme.com', 'aidshealth.org', 'asu.edu.eg', 'ait.global', 'pestcontrol.ae', 'openrainbow.com',
#  'alissoftware.com', 'altairengineering.it', 'altenew.com', 'amc-entertainment.com', 'americansystems.com', 'analyticalcannabis.com',
#  'ansys.com', 'apache.org', 'arcminnesota.org', 'arhs-group.com', 'arsgroup.com.ar', 'asgcorp.uk', 'assaabloy.com', 'att.com',
#  'atitesting.com', 'axiscapital.co.in', 'azzlee.com', 'aarauctions.com', 'abacustech.com', 'ru.abbott', 'abbts.ch', 'abercrombie.com',
#  'abilegroup.com', 'abodetechzone.com', 'acadaca.com', 'academy.com', 'acadiainsurance.com', 'atlab.com', 'accenture.com', 'acsgbl.com',
#  'aofund.org', 'acclaimtechnical.com', 'accordiluminacao.com', 'accuregsoftware.com', '4acetech.com', 'acm.org', 'acronis.com',
#  'actionet.com', 'activerelease.com', 'activesoft.com.br', 'activeyewear.com', 'actriv.com', 'actronsystems.com',
#  'acuityscheduling.com', 'acuitybrands.com', 'adsourced.com', 'adametrix.com', 'adamsservices.co.uk', 'adaptecsolutions.com']

# websites = ['adobe.com', 'apvit.com', 'aplacetocallhomeilp.org', 'aws-startup-lofts.com',
#          'copierserviceswestcovina.com', 'aaa.com', 'aaafoundationservice.com', 'aaratechnologies.com', 'aaratechnologies.com',
#          'aandbcleaningservices.co.uk', 'abc7.com', 'abctechnologies.com', 'abtassociates.com', 'ace-limited.info', 'aciinfotech.com',
#          'acvauctions.com', 'adbsafegate.com', 'ad.net', 'advantech.com', 'ae.com', 'aegworldwide.com', 'aesfluids.com',
#          'geodetics.com', 'ags.gob.mx', 'aheadofthyme.com', 'aidshealth.org', 'asu.edu.eg', 'ait.global', 'pestcontrol.ae',
#          'openrainbow.com', 'alissoftware.com', 'altairengineering.it', 'altenew.com', 'americansystems.com',
#          'analyticalcannabis.com', 'ansys.com', 'apache.org', 'arcminnesota.org', 'arhs-group.com', 'arsgroup.com.ar', 'asgcorp.uk',
#          'assaabloy.com', 'att.com', 'atitesting.com', 'axiscapital.co.in', 'azzlee.com', 'aarauctions.com', 'abacustech.com',
#          'ru.abbott', 'abbts.ch', 'abercrombie.com', 'abilegroup.com', 'abodetechzone.com', 'acadaca.com', 'academy.com',
#          'acadiainsurance.com', 'atlab.com', 'accenture.com', 'acsgbl.com', 'aofund.org', 'acclaimtechnical.com',
#          'accordiluminacao.com', 'accuregsoftware.com', 'acetechnologiesgroup.com', 'acm.org', 'acronis.com', 'actionet.com',
#          'activerelease.com', 'activesoft.com.br', 'activeyewear.com', 'actriv.com', 'actronsystems.com', 'acuityscheduling.com',
#          'acuitybrands.com', 'adsourced.com', 'adametrix.com', 'adamsservices.co.uk', 'adaptecsolutions.com']
        

from csv import writer
 

from datetime import date
SKILLS=['Developer', 'Engineer', 'Software Developer', 'Software Engineer', 'Junior Developer', 'Mid level developer',
        'Senior Developer', 'Programmer', 'Frontend developer', 'Frontend', 'Backend', 'Solutions Architect', 'Solution Architect'
        'Architect', 'Jr Developer','Ui/Ux', 'Designer', 'Coder', 'Entry Level', 'Fullstack', 'Full stack', 'Backend', 'laraval',
        'Php', 'quality assuarance', 'nodejs', '.Net', 'Cloud', 'DevOps', 'Python', 'Angular', 'Project Management', 'project manager'
        'junior developer','front-end', 'full-stack', ''
        ]
# file_path = "D:\\job source project\\JobManagement\\Skills.txt"
# with open(file_path, "r") as file:
#     SKILLS = file.read().splitlines()
    
# print(SKILLS)


# 
# SKILLS = ['full stack developer','backend developer', 'frontend developer','laraval','nodejs', 'dot net', 'fastapi', 'ui/ux','ui/ux designer', 'graphic designer', '.NET Core', 'ADO.net', 'Adobe Illustrator', 
#           'Adobe Photoshop', 'Amazon Web Services', 'Android App Development', 'Android developer', 'Android'
#           'Angular', 'Angular JS', 'Apache', 'Apache Cordova', 'Apex', 'APIs Testing', 'quality assuarance',  
#           'Application Architecture', 'ASP.NET', 'Solution Architect', 
#           'AWS Cloud9', 'Cloud developer', 'Cloud Architect',  
#           'Azure DevOps', 'BigQuery', 'Business Intelligence', 'DevOps Engineer',
#           'C#', 'C++', 'CodeIgniter', 'Salesforce developer','Salesforce developer', 
#         'Data Modeling', 'Data Warehousing', 'Database Administration', 'dba', 'Database Engineer', 
#           'Database Management', 'SQL Development', 'DAX Query', 
#           'DevOps', 'Django', 'Drupal',  
#            'ETL', 'Express.js', 'Figma',  'Flask', 
#           'Google Cloud', 'GraphQL','HubSpot', 'Java', 'Java Swing', 
#           'Jmeter', 'Knockout.js', 
#           'Kubernetes', 'Lambda', 'Laravel', 'Linux Admin', 
#           'Magento', 'Microsoft Azure', 
#           'MongoDB', 'MS SQL Server', 'MySQL', 'Nest.js',
#           'Next.js', 'Nginx', 'Node.js', 'NoSQL Databases', 'Oracle developer','Oracle database', 
#           'PHP', 'PostgreSQL', 'Power BI', 
#         'Project Management', 'project manager' 
           
#            'Python', 'Quality Control', 
#           'React Native', 'ReactJS', 'Redis Server', 'Redux', 'RESTful API', 'Ruby', 
#            'SASS', 'Scrum master', 'Perl', 'developer' 
#           'SCSS', 'Selenium', 'Shopify', 
#            'SQL Server', 
#           'Tailwind CSS', 'Tailwind', 
#           'Test Automation', 'Test Cases', 
#           'Twilio', 'TypeScript', 'Ubuntu Linux',  
#          'Vue.js developer', 'Web APIs', 'Web Service API',  
#           'WooCommerce', 'WordPress' ]

# SKILLS = [  '.NET Core', 'Active Directory', 'ADO.net', 'Adobe Illustrator', 
#           'Adobe Photoshop', 'Agile methodology', 'AJAX', 'Amazon Web Services', 'Android App Development', 
#           'Angular', 'Angular JS', 'Ansible', 'Ant Design', 'Apache', 'Apache Cordova', 'Apex', 'APIs Testing', 
#           'Application Architecture', 'Apttus Contract Management - Salesforce', 'ASP.NET', 'ASP.NET Identity', 'ASP.NET MVC', 
#           'ASP.NET SignalR', 'ASP.NET Web API', 'ASP.NET Web Forms', 'AWS Cloud9', 'Axios', 'Azure Administration', 
#           'Azure DevOps Server', 'Bash Scripting', 'BigQuery', 'Blazor', 'Bootstrap', 'Business Intelligence', 
#           'C#', 'C++', 'CodeIgniter', 'ColdFusion', 'Copado DevOps - Salesforce', 'Crystal Reports', 'CSS', 
#           'Cypress.io', 'Dapper', 'Data Dictionary', 'Data Modeling', 'Data Warehousing', 'Database Administration', 
#           'Database Management', 'Database/SQL Development', 'DataLoader - Salesforce', 'DAX Query', 'Demand Tool - Salesforce', 
#           'DevOps Engineering', 'Django', 'DNS Servers', 'Docker', 'Dreamweaver', 'Drupal', 'EC2', 'Electron', 
#           'Entity Framework', 'ETL', 'Express.js', 'Facebook API', 'Figma', 'Firebase', 'Flask', 'Git', 
#           'Google Cloud', 'Google Maps', 'GraphQL', 'HTML', 'HubSpot', 'Informatica', 'Java', 'Java Swing', 
#           'JavaScript', 'JIRA', 'Jmeter (Apache load testing)', 'jQuery', 'jQuery Mobile', 'Knockout.js', 
#           'Kubernetes', 'Lambda', 'Laravel', 'Lightning Component', 'Lightning Web Component', 'LINQ', 'Linux', 
#           'Magento', 'Manual Testing Types', 'Microsoft Azure', 'Microsoft Exchange', 'Microsoft Power Apps', 
#           'Mobile Application Testing', 'MongoDB', 'MS SQL Server', 'MySQL', 'Nest.js', 'NetBeans', 'Network Troubleshooting', 
#           'Next.js', 'Nginx', 'Node.js', 'NoSQL Databases', 'Office 365', 'OpenCart', 'Oracle', 'Oracle Forms', 
#           'OWASP Top Ten', 'Phalcon', 'PHP', 'PL/SQL', 'PostgreSQL', 'Power BI', 'PowerAutomate', 
#           'Project Communication Management', 'Project Human Resource Management', 'Project Integration Management', 'Project Management', 
#           'Project Procurement Management', 'Project Quality Management', 'Project Risk Management', 'Project Schedule Management', 
#           'Project Scope Management', 'Project Stakeholder Management', 'Python', 'Quality Control Testing', 
#           'Razor Page', 'RDS', 'React Native', 'ReactJS', 'Redis Server', 'Redux', 'RESTful API', 'Ruby on Rails', 
#           'S3 Storage', 'Salesforce', 'Salesforce Administration', 'Salesforce Process Builder', 'Salesforce Sales Cloud', 
#           'Salesforce Service Cloud', 'Salesforce Validation Rules', 'Salesforce Workflow Rules', 'SASS', 'Scrum methodology', 
#           'SCSS', 'SDLC Models', 'Selenium', 'Send Grid API', 'SharePoint', 'Shopify API', 'Slack API', 
#           'Socket.IO', 'SOQL - Salesforce', 'SOSL - Salesforce', 'SQL Server Analysis Services (SSAS)', 
#           'SSH', 'SSIS', 'SSRS', 'Sugar CRM', 'SVN', 'T-SQL', 'Tableau', 'Tailwind CSS Framework', 
#           'Team Foundation Server (TFS)', 'Telerik Controls', 'Test Automation', 'Test Cases', 'Test Plans', 
#           'Test Strategy', 'Testing Web Apps', 'Twilio', 'Twitter API', 'TypeScript', 'Ubuntu Linux', 'Visual Studio', 
#           'VisualForce Pages', 'VMware Virtualization', 'Vue.js', 'Web APIs', 'Web Service API', 'Windows Servers', 
#           'WooCommerce', 'WordPress', 'WordPress Plugins', 'WordPress Theme Development', 'XML', 'Youtube API']
from django.http import HttpResponse
def ajaxcall(request):
    
    websites = ['montana.edu', 'stripe.com', '1800flowers.com', 'bitsgap.com', 'pointclickcare.com', 'brothers-brick.com', 
'sevenrooms.com', '2ndswing.com', '2ndwatch.com', '3boxlabs.com', '3dsystems.com', 'coresystems.hu', '3i-infotech.com',
'4cgeoworks.com', 'patriotsoftware.com', '32degrees.com', 'kingautomotive.net', 'elevenwarriors.com', 'facebook.com',
'summitsalon.com', '17lumberandrental.com', 'xhamster.com', 'adobe.com', 'apvit.com', 'aplacetocallhomeilp.org', 
'aws-startup-lofts.com', 'copierserviceswestcovina.com', 'aaa.com', 'aaafoundationservice.com', 'aaratechnologies.com', 
'aaratechnologies.com', 'aandbcleaningservices.co.uk', 'abc7.com', 'abctechnologies.com', 'abtassociates.com', 'ace-limited.info', 
'aciinfotech.com', 'acvauctions.com', 'adbsafegate.com', 'ad.net', 'advantech.com', 'ae.com', 'aegworldwide.com', 'aesfluids.com',
'geodetics.com', 'ags.gob.mx', 'aheadofthyme.com', 'aidshealth.org', 'asu.edu.eg', 'ait.global', 'pestcontrol.ae', 'openrainbow.com',
'alissoftware.com', 'altairengineering.it', 'altenew.com', 'amc-entertainment.com', 'americansystems.com', 'analyticalcannabis.com',
'ansys.com', 'apache.org', 'arcminnesota.org', 'arhs-group.com', 'arsgroup.com.ar', 'asgcorp.uk', 'assaabloy.com', 'att.com',
'atitesting.com', 'axiscapital.co.in', 'azzlee.com', 'aarauctions.com', 'abacustech.com', 'ru.abbott', 'abbts.ch', 'abercrombie.com',
'abilegroup.com', 'abodetechzone.com', 'acadaca.com', 'academy.com', 'acadiainsurance.com', 'atlab.com', 'accenture.com', 'acsgbl.com',
'aofund.org', 'acclaimtechnical.com', 'accordiluminacao.com', 'accuregsoftware.com', '4acetech.com', 'acm.org', 'acronis.com',
'actionet.com', 'activerelease.com', 'activesoft.com.br', 'activeyewear.com', 'actriv.com', 'actronsystems.com',
'acuityscheduling.com', 'acuitybrands.com', 'adsourced.com', 'adametrix.com', 'adamsservices.co.uk', 'adaptecsolutions.com']
#    'actronsystems.com',
    # websites = ['accenture.com']
    # websites = ['montana.edu', 'stripe.com','acuitybrands.com', 'adsourced.com', 'adametrix.com', 'adamsservices.co.uk', 'adaptecsolutions.com']
    # print(websites)
    # print(len(websites))
    
    for result in websites:
    # if result:
        try:
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
            # final = "https://" + result[0]['domain']
            # final = "http://" + result[0]['domain']
            final = "http://" + result
            
            # final = "https://https://talentservice.com/" 
            driver.get(final)
            # print('.........................................................................')
            try:
                data = []
                data_with_company = []
                # TODO fetch keyword from csv file (make csv file) -> today
                keywords = ['Careers','Career', 'Employment', 'Job', 'Jobs', 'Apply Now']
                for keyword in keywords:
                    try:
                        # element = driver.find_element(By.XPATH, "//a[contains(text(), 'Career')]") worked for acadaca
                        element = driver.find_element(By.XPATH, f"//*[contains(text(), '{keyword}')]") 
                        # print(element.tag_name)
                        # time.sleep(3)
                        WebDriverWait(driver, 1).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')
                        try:
                            element = getElementByTag(element, 'a')
                        except Exception as e:
                            element = None
                            # print(e)
                        if element:
                            if element.is_enabled():
                                # ActionChains(driver).move_to_element(element).perform()
                                # print("Keyword found and clickable!")
                                # print(element.get_attribute("innerHTML"))
                                # driver.get(element.get_attribute('href'))
                                
                                # print(element.get_attribute('href'))
                                # print("Element is visible? " + str(element.is_displayed()))
                                # print(element.tag_name)
                                driver.execute_script("arguments[0].click();", element)
                                # ActionChains(driver).move_to_element(element).perform()
                                # element.send_keys(Keys.RETURN) 
                                WebDriverWait(driver, 1).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')
                                
                                # time.sleep(3)
                                found = True
                                # 2nddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddd pageeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee
                                # print('**********************************************************')
                                get_url = driver.current_url
                                page_source = driver.page_source
                                # print(get_url)
                                # el = driver.execute_script("return $('*:contains('AWS')'); ")
                                # print('**********************************************************')
                                second_page_keywords = ['OPEN POSITIONS', 'Search Jobs', 'Find a Career','See Jobs','open roles', 'Open Opportunities' ,'open jobs', 'CURRENT OPENINGS','Join Us', 'Information Technology']
                                for second_keyword in second_page_keywords:
                                    try:
                                        # print('2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd')
                                        second_page_element = driver.find_element(By.XPATH, f"//*[contains(text(), '{second_keyword}')]")
                                        # second_page_element = driver.find_element(By.XPATH, f"//a[contains(text(), '{second_keyword}')]")
                                        print(second_page_element.get_attribute("innerHTML"))
#uncomment this                                           # element = getElementByTag(second_page_element, 'a')
#uncomment this                                           # if element.is_enabled():
                                        if second_page_element.is_enabled():
                                            
                                            print('ENABLEDDDDDDDDDDDDDDDDDDDDDDDDDDDDDDDD')
                                            # print(element.get_attribute("innerHTML"))
                                            # print(element.get_attribute('href'))
                                            # print("Element is visible? " + str(element.is_displayed()))
                                            WebDriverWait(driver, 2).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')
                                            
                                            # time.sleep(2)
                                            print('ENABLEDDDDDDDDDDDDDDDDDDDDDDDDDDDDDDDD')
                                            
                                            driver.execute_script("arguments[0].click();", element)
                                            
                                    except Exception as e:
                                        print(e)
                                        pass
                                WebDriverWait(driver, 1).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')
                                
                                # time.sleep(10)
                                # while driver.find_element(By.XPATH, f"//a[contains(text(), '{second_keyword}')]")
                                next_element =False
                                next_page = True
                                
                                while next_page:
                                    previous =driver.current_url
                                    # file_object = open('metasource.txt', 'a')
                                    # file_object.write(driver.page_source)
                                    # file_object.close()
                                    try:
                                        next_element = driver.find_element(By.XPATH, "//*[contains(text(), 'Next')]")
                                        # print(next_element.tag_name)
                                    except Exception as e:
                                        next_element = False
                                        # print(e)
                                #     # data = getJobsfromCurrentPage(driver, data)
                                #     # data.append(current_jobs_data)

                                    print('00000000000000000000000000000000000000000000000000000')
                                    # print('00000000000000000000000000000000000000000000000000000')
                                    # print('00000000000000000000000000000000000000000000000000000')
                                    # print(driver.page_source)
                                    # print(driver.current_url)
                                    
                                    # time.sleep(10)
                                    # try:
                                    #     with open('meta.txt', "w", encoding="utf-8") as f:
                                    #         f.write(driver.page_source)
                                    # except Exception as e:
                                        # print('----------------------------')
                                        # print(e)
                                        # print('----------------------------')
                                        
                                    # print('dumpped')
                                    
                                    for skill in SKILLS:
                                        # remove tag
                                        # skill_xpath = f"//body[contains(text(), '{skill}')]"
                                        
                                        # skill_xpath = f"//*[contains(text(), '{skill}')]"
                                        skill_xpath = f"/html/body//*[contains(text(), '{skill}')]"
                                        
                                        # print(skill_xpath)
                                        # print('skill loop starting')
                                        try:
                                            # print(driver.page_source)
                                            
                                            element = driver.find_element(By.XPATH, skill_xpath) 

                                            if element: 
                                                # print('skill is there')
                                                # print('Tag Name   Tag Name   Tag Name   Tag Name   Tag Name   Tag Name   ')
                                                # print('Tag Name   Tag Name   Tag Name   Tag Name   Tag Name   Tag Name   ')
                                                
                                                # print(element.tag_name)
                                                # print(element.get_attribute("innerHTML"))
                                                
                                                try:
                                                    element_link = getElementByTag(element, 'a')
                                                except Exception as e:
                                                    element_link=None
                                                    # print(e)
                                                # try:
                                                #     print('beforte click beforte click beforte click beforte click beforte click ')
                                                #     # element.send_keys(Keys.COMMAND + 't') 
                                                #     driver.execute_script("arguments[0].click();", element)
                                                    
                                                #     print(skill_xpath)
                                                #     print(element.get_attribute("innerHTML"))
                                                #     time.sleep(12)
                                                    
                                                # except Exception as e:
                                                #     print('I am unable to click so dont click me')
                                                #     print('I am unable to click so dont click me')
                                                #     print('I am unable to click so dont click me')
                                                # time.sleep(10)
                                                # print('awaken')
                                                # print('awaken')
                                                print('awaken')
                                                
                                                if element_link:
                                                    # print('element_link')
                                                    # print(element_link)
                                                    # element_link.get_attribute('href')
                                                    
                                                    # print(element.get_attribute("innerHTML"))
                                                    # d = {"Skill": skill,"Job Title" : element.get_attribute("innerHTML"), "Job Link": element_link.get_attribute('href')}
                                                    d = {"Skill": skill,"Job Title" : element_link.get_attribute("innerHTML"), 
                                                            "Job Link": element_link.get_attribute('href'),
                                                            "Source": previous}
                                                    data.append(d)
                                                    f = {"Company Site": result,"Skill": skill,"Job Title" : element_link.get_attribute("innerHTML"), 
                                                            "Job Link": element_link.get_attribute('href'),
                                                            "Source": previous}
                                                    data_with_company.append(f)
                                                    
                                                # elif element:
                                                #     # print('innerHTML')
                                                #     # print(element.get_attribute("innerHTML"))
                                                #     d = {
                                                #         "Skill": skill,"Job Title" : element.get_attribute("innerHTML"), 
                                                #          "Job Link": "Link Not Available",
                                                #          "Source": previous
                                                #         }
                                                #     data.append(d)
                                                #     f = {"Company Site": result,"Skill": skill,"Job Title" : element.get_attribute("innerHTML"), 
                                                #          "Job Link": "Link Not Available",
                                                #          "Source": previous
                                                #          }
                                                #     data_with_company.append(f)
                                                    
                                                    
                                        except Exception as e:
                                            # print('in except in except')
                                            # print(e)
                                            pass
                                    # print(next_element)
                                        
                                    if not next_element:
                                        # print('not in next_element')
                                        next_page = False
                                    elif next_element:
                                        print(' in in in next_page')
                                        
                                        if  next_element.is_enabled():
                                            
                                            # print('Clicking Clicking Clicking Clicking Clicking Clicking Clicking ')
                                            # next_element.click()
                                            driver.execute_script("arguments[0].click();", next_element)
                                            # time.sleep(2)
                                            WebDriverWait(driver, 1).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')
                                            
                                            # time.sleep(5)
                                            
                                            current = driver.current_url
                                            
                                            if previous == current:
                                                next_page = False
                                        else:
                                            # print('before while wlseeeeeeeeeeeeeeeeeeee')
                                            next_page = False
                                # print(data)
                                # print(data_with_company)
                                # print('after while++++++++++++++++')
                                # print(data)
                                if data:     
                                    saveToExcelFile(data, result)
                                    saveTAllDataExcel(data_with_company, result)
                                    
                                    # saveToCsvFile(data, result)
                                    # saveTAllDataCsv(data_with_company, result)
                                    #     writer.writeheader()

                                    #     # Write the data rows
                                    #     writer.writerows(data)
                                    # print('Data has been successfully written to', csv_file)
                                    
                                            
                                # driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                # driver.implicitly_wait(3) 
                                # next_element = driver.find_element(By.XPATH, "//*[contains(text(), 'FIND YOUR JOB')]")
                                # if next_element:
                                #     next_element.send_keys(Keys.RETURN) 
                                    #  elementz = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, class_name)))
                                    # //div[@class='curJobTitle']/a
                                # driver.execute_script("window.stop();")
                        # for link in links:
                        #     if link.get_attribute("innerHTML") == "Careers":
                        #         print(link.get_attribute("innerHTML"))
                        #         link.click()
                                # wait = WebDriverWait(driver, 10)
                                # wait.until
                        #         found =True
                        #         break
                        # if found:
                        #     # new_element = driver.find_element(By.XPATH, "//a[@class='job_title_link']")
                        #     # new_element = driver.find_elements(By.XPATH, "//div[@class='curJobTitle']")
                        #     new_element = driver.find_elements("xpath", "//p/span") 
                        #     # new_element = driver.find_elements("xpath", "//p/span/a") #for abacustech
                            # xpath= "//div[@class='cmp-teaser__content']/h3" accenture
                            # new_element = driver.find_elements("xpath", "//div[@class='curJobTitle']/a") acsgbl
                            # new_element = driver.find_elements("xpath", "//div[@id='featured-job-listing']/ul/li/a/h3")
                            # field_names = ['Skill', 'Job Title', 'Job Link']
                                
                            # csv_file = f"{company_name}.csv"
                            # with open(csv_file, 'w', newline='') as file:
                            #     writer = csv.DictWriter(file, fieldnames=field_names)

                            #     # Write the header row
                            #     writer.writeheader()

                            #     # Write the data rows
                            #     writer.writerows(data)
                            # print('Data has been successfully written to', csv_file)
                            # element = driver.find_element(By.XPATH, "//a[contains(text(), 'SEARCH JOBS')]")
                            # if element.is_enabled():
                            #     element.click()
                    except Exception as e:
                        # print(e)
                        # print("Keyword after exception")
                        pass
                        # link2 = driver.find_elements("xpath", "//a[contains(text(), 'SEARCH JOBS')]")
                        # print(link2)
                        
                        # for l in link2:
                        #     print(l)
                driver.quit()
            except Exception as e:
                driver.quit()            
            # driver.maximize_window()
            
            # Access Global
        except Exception as e:
            # print(e)
            pass
        
        
        # result = json.loads(result)
        # print(result[0]["domain"])
    # return render(request, 'demo/form.html')
    return HttpResponse("Raza")
def index(request):
                    
    if request.method == 'POST':
        # SKILLS = fetchSkills()
        # print(SKILLS)
        # df = pd.read_excel(company_name, usecols='A')
        # print(df.columns)
        # for column in df.columns:
        #     list1 = df[column].unique().tolist()
        # print(list1)
        # print(len(list1))
        try:
            company_name = request.FILES['company']
        except MultiValueDictKeyError:
            return render(request, 'demo/form.html')
        splitting = str(company_name).split('.')[-1]
        
        if (splitting == 'xls'):
            data = xls_get(company_name, column=2)
        elif (splitting == 'xlsx'):
            data = xlsx_get(company_name, column=2)
        else:
            return render(request, 'demo/form.html')
        sheet = 0
        for i in data.keys():
            sheet = i
            break
        companies = data[sheet]
        job_titles = []
        websites = []
        websites_not_found = []
        # https://www.aciinfotech.com/ issue with it
        index=0
        # for company in companies[1:]:
        #     if company[1] == "Direct Employer" :
        #         x = re.sub("\" | Inc.| Llc| LLC|\'|, LLC|, Inc.|, Inc.|, Llc| Inc|, Inc| Inc.| Inc| INC|, LLC.|, Ltd|\.|,","", company[0])
        #         if x not in job_titles: 
        #             job_titles.append(x)
                    
        #             # print(x)
        #             url = "https://autocomplete.clearbit.com/v1/companies/suggest?query="+x
        #             index+=1
        #             # print(index)
        #             result = requests.get(url)
        #             result = result.json()
        #             if result:
        #                 websites.append(result[0]['domain'])
        #             else:
        #                 websites_not_found.append(x)
                        # till
        # company_name = request.POST['company_title']
        # print(company_name)
        # url = "https://autocomplete.clearbit.com/v1/companies/suggest?query="+company_name
        # result = requests.get(url)
        # result = result.json()
        try:
            # pass
            os.remove('CompleteData.xlsx')    
        except:
            pass
        # websites = ['3boxlabs.com']
        websites = ['montana.edu', 'stripe.com', '1800flowers.com', 'bitsgap.com', 'pointclickcare.com', 'brothers-brick.com', 
 'sevenrooms.com', '2ndswing.com', '2ndwatch.com', '3boxlabs.com', '3dsystems.com', 'coresystems.hu', '3i-infotech.com',
 '4cgeoworks.com', 'patriotsoftware.com', '32degrees.com', 'kingautomotive.net', 'elevenwarriors.com', 'facebook.com',
 'summitsalon.com', '17lumberandrental.com', 'xhamster.com', 'adobe.com', 'apvit.com', 'aplacetocallhomeilp.org', 
 'aws-startup-lofts.com', 'copierserviceswestcovina.com', 'aaa.com', 'aaafoundationservice.com', 'aaratechnologies.com', 
 'aaratechnologies.com', 'aandbcleaningservices.co.uk', 'abc7.com', 'abctechnologies.com', 'abtassociates.com', 'ace-limited.info', 
 'aciinfotech.com', 'acvauctions.com', 'adbsafegate.com', 'ad.net', 'advantech.com', 'ae.com', 'aegworldwide.com', 'aesfluids.com',
 'geodetics.com', 'ags.gob.mx', 'aheadofthyme.com', 'aidshealth.org', 'asu.edu.eg', 'ait.global', 'pestcontrol.ae', 'openrainbow.com',
 'alissoftware.com', 'altairengineering.it', 'altenew.com', 'amc-entertainment.com', 'americansystems.com', 'analyticalcannabis.com',
 'ansys.com', 'apache.org', 'arcminnesota.org', 'arhs-group.com', 'arsgroup.com.ar', 'asgcorp.uk', 'assaabloy.com', 'att.com',
 'atitesting.com', 'axiscapital.co.in', 'azzlee.com', 'aarauctions.com', 'abacustech.com', 'ru.abbott', 'abbts.ch', 'abercrombie.com',
 'abilegroup.com', 'abodetechzone.com', 'acadaca.com', 'academy.com', 'acadiainsurance.com', 'atlab.com', 'accenture.com', 'acsgbl.com',
 'aofund.org', 'acclaimtechnical.com', 'accordiluminacao.com', 'accuregsoftware.com', '4acetech.com', 'acm.org', 'acronis.com',
 'actionet.com', 'activerelease.com', 'activesoft.com.br', 'activeyewear.com', 'actriv.com', 'actronsystems.com',
 'acuityscheduling.com', 'acuitybrands.com', 'adsourced.com', 'adametrix.com', 'adamsservices.co.uk', 'adaptecsolutions.com']
    #    'actronsystems.com',
        # websites = ['accenture.com']
        # websites = ['montana.edu', 'stripe.com','acuitybrands.com', 'adsourced.com', 'adametrix.com', 'adamsservices.co.uk', 'adaptecsolutions.com']
        # print(websites)
        # print(len(websites))
        
        for result in websites:
        # if result:
            try:
                driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
                # final = "https://" + result[0]['domain']
                # final = "http://" + result[0]['domain']
                final = "http://" + result
                
                # final = "https://https://talentservice.com/" 
                driver.get(final)
                # print('.........................................................................')
                try:
                    data = []
                    data_with_company = []
                    # TODO fetch keyword from csv file (make csv file) -> today
                    keywords = ['Careers','Career', 'Employment', 'Job', 'Jobs', 'Apply Now']
                    for keyword in keywords:
                        try:
                            # element = driver.find_element(By.XPATH, "//a[contains(text(), 'Career')]") worked for acadaca
                            element = driver.find_element(By.XPATH, f"//*[contains(text(), '{keyword}')]") 
                            # print(element.tag_name)
                            # time.sleep(3)
                            WebDriverWait(driver, 1).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')
                            try:
                                element = getElementByTag(element, 'a')
                            except Exception as e:
                                element = None
                                # print(e)
                            if element:
                                if element.is_enabled():
                                    # ActionChains(driver).move_to_element(element).perform()
                                    # print("Keyword found and clickable!")
                                    # print(element.get_attribute("innerHTML"))
                                    # driver.get(element.get_attribute('href'))
                                    
                                    # print(element.get_attribute('href'))
                                    # print("Element is visible? " + str(element.is_displayed()))
                                    # print(element.tag_name)
                                    driver.execute_script("arguments[0].click();", element)
                                    # ActionChains(driver).move_to_element(element).perform()
                                    # element.send_keys(Keys.RETURN) 
                                    WebDriverWait(driver, 1).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')
                                    
                                    # time.sleep(3)
                                    found = True
                                    # 2nddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddd pageeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee
                                    # print('**********************************************************')
                                    get_url = driver.current_url
                                    page_source = driver.page_source
                                    # print(get_url)
                                    # el = driver.execute_script("return $('*:contains('AWS')'); ")
                                    # print('**********************************************************')
                                    second_page_keywords = ['OPEN POSITIONS', 'Search Jobs', 'Find a Career','See Jobs','open roles', 'Open Opportunities' ,'open jobs', 'CURRENT OPENINGS','Join Us', 'Information Technology']
                                    for second_keyword in second_page_keywords:
                                        try:
                                            # print('2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd 2nd')
                                            second_page_element = driver.find_element(By.XPATH, f"//*[contains(text(), '{second_keyword}')]")
                                            # second_page_element = driver.find_element(By.XPATH, f"//a[contains(text(), '{second_keyword}')]")
                                            print(second_page_element.get_attribute("innerHTML"))
 #uncomment this                                           # element = getElementByTag(second_page_element, 'a')
   #uncomment this                                           # if element.is_enabled():
                                            if second_page_element.is_enabled():
                                                
                                                print('ENABLEDDDDDDDDDDDDDDDDDDDDDDDDDDDDDDDD')
                                                # print(element.get_attribute("innerHTML"))
                                                # print(element.get_attribute('href'))
                                                # print("Element is visible? " + str(element.is_displayed()))
                                                WebDriverWait(driver, 2).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')
                                                
                                                # time.sleep(2)
                                                print('ENABLEDDDDDDDDDDDDDDDDDDDDDDDDDDDDDDDD')
                                                
                                                driver.execute_script("arguments[0].click();", element)
                                                
                                        except Exception as e:
                                            print(e)
                                            pass
                                    WebDriverWait(driver, 1).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')
                                    
                                    # time.sleep(10)
                                    # while driver.find_element(By.XPATH, f"//a[contains(text(), '{second_keyword}')]")
                                    next_element =False
                                    next_page = True
                                    
                                    while next_page:
                                        previous =driver.current_url
                                        # file_object = open('metasource.txt', 'a')
                                        # file_object.write(driver.page_source)
                                        # file_object.close()
                                        try:
                                            next_element = driver.find_element(By.XPATH, "//*[contains(text(), 'Next')]")
                                            # print(next_element.tag_name)
                                        except Exception as e:
                                            next_element = False
                                            # print(e)
                                    #     # data = getJobsfromCurrentPage(driver, data)
                                    #     # data.append(current_jobs_data)

                                        print('00000000000000000000000000000000000000000000000000000')
                                        # print('00000000000000000000000000000000000000000000000000000')
                                        # print('00000000000000000000000000000000000000000000000000000')
                                        # print(driver.page_source)
                                        # print(driver.current_url)
                                        
                                        # time.sleep(10)
                                        # try:
                                        #     with open('meta.txt', "w", encoding="utf-8") as f:
                                        #         f.write(driver.page_source)
                                        # except Exception as e:
                                            # print('----------------------------')
                                            # print(e)
                                            # print('----------------------------')
                                            
                                        # print('dumpped')
                                        
                                        for skill in SKILLS:
                                            # remove tag
                                            # skill_xpath = f"//body[contains(text(), '{skill}')]"
                                            
                                            # skill_xpath = f"//*[contains(text(), '{skill}')]"
                                            skill_xpath = f"/html/body//*[contains(text(), '{skill}')]"
                                            
                                            # print(skill_xpath)
                                            # print('skill loop starting')
                                            try:
                                                # print(driver.page_source)
                                                
                                                element = driver.find_element(By.XPATH, skill_xpath) 

                                                if element: 
                                                    # print('skill is there')
                                                    # print('Tag Name   Tag Name   Tag Name   Tag Name   Tag Name   Tag Name   ')
                                                    # print('Tag Name   Tag Name   Tag Name   Tag Name   Tag Name   Tag Name   ')
                                                    
                                                    # print(element.tag_name)
                                                    # print(element.get_attribute("innerHTML"))
                                                    
                                                    try:
                                                        element_link = getElementByTag(element, 'a')
                                                    except Exception as e:
                                                        element_link=None
                                                        # print(e)
                                                    # try:
                                                    #     print('beforte click beforte click beforte click beforte click beforte click ')
                                                    #     # element.send_keys(Keys.COMMAND + 't') 
                                                    #     driver.execute_script("arguments[0].click();", element)
                                                        
                                                    #     print(skill_xpath)
                                                    #     print(element.get_attribute("innerHTML"))
                                                    #     time.sleep(12)
                                                        
                                                    # except Exception as e:
                                                    #     print('I am unable to click so dont click me')
                                                    #     print('I am unable to click so dont click me')
                                                    #     print('I am unable to click so dont click me')
                                                    # time.sleep(10)
                                                    # print('awaken')
                                                    # print('awaken')
                                                    print('awaken')
                                                    
                                                    if element_link:
                                                        # print('element_link')
                                                        # print(element_link)
                                                        # element_link.get_attribute('href')
                                                        
                                                        # print(element.get_attribute("innerHTML"))
                                                        # d = {"Skill": skill,"Job Title" : element.get_attribute("innerHTML"), "Job Link": element_link.get_attribute('href')}
                                                        d = {"Skill": skill,"Job Title" : element_link.get_attribute("innerHTML"), 
                                                             "Job Link": element_link.get_attribute('href'),
                                                             "Source": previous}
                                                        data.append(d)
                                                        f = {"Company Site": result,"Skill": skill,"Job Title" : element_link.get_attribute("innerHTML"), 
                                                             "Job Link": element_link.get_attribute('href'),
                                                             "Source": previous}
                                                        data_with_company.append(f)
                                                        
                                                    # elif element:
                                                    #     # print('innerHTML')
                                                    #     # print(element.get_attribute("innerHTML"))
                                                    #     d = {
                                                    #         "Skill": skill,"Job Title" : element.get_attribute("innerHTML"), 
                                                    #          "Job Link": "Link Not Available",
                                                    #          "Source": previous
                                                    #         }
                                                    #     data.append(d)
                                                    #     f = {"Company Site": result,"Skill": skill,"Job Title" : element.get_attribute("innerHTML"), 
                                                    #          "Job Link": "Link Not Available",
                                                    #          "Source": previous
                                                    #          }
                                                    #     data_with_company.append(f)
                                                        
                                                        
                                            except Exception as e:
                                                # print('in except in except')
                                                # print(e)
                                                pass
                                        # print(next_element)
                                            
                                        if not next_element:
                                            # print('not in next_element')
                                            next_page = False
                                        elif next_element:
                                            print(' in in in next_page')
                                            
                                            if  next_element.is_enabled():
                                                
                                                # print('Clicking Clicking Clicking Clicking Clicking Clicking Clicking ')
                                                # next_element.click()
                                                driver.execute_script("arguments[0].click();", next_element)
                                                # time.sleep(2)
                                                WebDriverWait(driver, 1).until(lambda driver: driver.execute_script('return document.readyState') == 'complete')
                                                
                                                # time.sleep(5)
                                                
                                                current = driver.current_url
                                                
                                                if previous == current:
                                                    next_page = False
                                            else:
                                                # print('before while wlseeeeeeeeeeeeeeeeeeee')
                                                next_page = False
                                    # print(data)
                                    # print(data_with_company)
                                    # print('after while++++++++++++++++')
                                    # print(data)
                                    if data:     
                                        saveToExcelFile(data, result)
                                        saveTAllDataExcel(data_with_company, result)
                                        
                                        # saveToCsvFile(data, result)
                                        # saveTAllDataCsv(data_with_company, result)
                                        #     writer.writeheader()

                                        #     # Write the data rows
                                        #     writer.writerows(data)
                                        # print('Data has been successfully written to', csv_file)
                                        
                                                
                                    # driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                    # driver.implicitly_wait(3) 
                                    # next_element = driver.find_element(By.XPATH, "//*[contains(text(), 'FIND YOUR JOB')]")
                                    # if next_element:
                                    #     next_element.send_keys(Keys.RETURN) 
                                        #  elementz = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, class_name)))
                                        # //div[@class='curJobTitle']/a
                                    # driver.execute_script("window.stop();")
                            # for link in links:
                            #     if link.get_attribute("innerHTML") == "Careers":
                            #         print(link.get_attribute("innerHTML"))
                            #         link.click()
                                    # wait = WebDriverWait(driver, 10)
                                    # wait.until
                            #         found =True
                            #         break
                            # if found:
                            #     # new_element = driver.find_element(By.XPATH, "//a[@class='job_title_link']")
                            #     # new_element = driver.find_elements(By.XPATH, "//div[@class='curJobTitle']")
                            #     new_element = driver.find_elements("xpath", "//p/span") 
                            #     # new_element = driver.find_elements("xpath", "//p/span/a") #for abacustech
                                # xpath= "//div[@class='cmp-teaser__content']/h3" accenture
                                # new_element = driver.find_elements("xpath", "//div[@class='curJobTitle']/a") acsgbl
                                # new_element = driver.find_elements("xpath", "//div[@id='featured-job-listing']/ul/li/a/h3")
                                # field_names = ['Skill', 'Job Title', 'Job Link']
                                    
                                # csv_file = f"{company_name}.csv"
                                # with open(csv_file, 'w', newline='') as file:
                                #     writer = csv.DictWriter(file, fieldnames=field_names)

                                #     # Write the header row
                                #     writer.writeheader()

                                #     # Write the data rows
                                #     writer.writerows(data)
                                # print('Data has been successfully written to', csv_file)
                                # element = driver.find_element(By.XPATH, "//a[contains(text(), 'SEARCH JOBS')]")
                                # if element.is_enabled():
                                #     element.click()
                        except Exception as e:
                            # print(e)
                            # print("Keyword after exception")
                            pass
                            # link2 = driver.find_elements("xpath", "//a[contains(text(), 'SEARCH JOBS')]")
                            # print(link2)
                            
                            # for l in link2:
                            #     print(l)
                    driver.quit()
                except Exception as e:
                    driver.quit()            
                # driver.maximize_window()
                
                # Access Global
            except Exception as e:
                # print(e)
                pass
            
        
        # result = json.loads(result)
        # print(result[0]["domain"])
    return render(request, 'demo/form.html')
# Uploading and refinement of  
# scrapping & fetching of 
#   

# def saveTAllDataCsv(data, result):
#     # field_names = ['Company Site','Skill', 'Job Title', 'Job Link']
#     # csv_file = "CompleteData.csv"
#     try:
#         excel_file = f"{result.split('.')[0]}"
#         wb = openpyxl.Workbook()   
#         must_add_headers = False if os.path.isfile(csv_file) else True
#         with open(csv_file, 'a', newline='') as file:
#             writer = csv.DictWriter(file, fieldnames=field_names)
#             if must_add_headers:
#                 writer.writeheader()
#             writer.writerows(data)
                
#         print('Data has been successfully written to', csv_file)
#     except Exception as e:
#         # print(e)
#         pass 


def saveToExcelFile(data, result):
    try:
        today = str(date.today())
        excel_file = f"{result.split('.')[0]}"
        wb = openpyxl.Workbook() 
        # Sheet_name = wb.sheetnames
        ws = wb.active
        ws.append(
            ('Skill','Job Title','Job Link', 'Source')
        )
        for i in data: 
            j =(
                i['Skill'],i['Job Title'],i['Job Link'], i['Source'] 
            )  
            ws.append(j)
            
        wb.save(f'{excel_file}-{today}.xlsx')
    except Exception as e:
        pass
        
        
# field_names = ['Company Site','Skill', 'Job Title', 'Job Link']
#     try:  
#         must_add_headers = False if os.path.isfile(excel_file) else True
#         with open(excel_file, 'a', newline='') as file:
#             writer = csv.DictWriter(file, fieldnames=field_names)
#             if must_add_headers:
#                 writer.writeheader()
#             writer.writerows(data)
                
#         print('Data has been successfully written to', excel_file)
#     except Exception as e:
#         # print(e)
#         pass
from openpyxl import load_workbook
def saveTAllDataExcel(data, result):
    try:
        today = str(date.today())
        # excel_file = f"{result.split('.')[0]}"
        # excel_file = "CompleteData.csv"
        
        excel_file = f"CompleteData-{today}.xlsx"
        must_add_headers = False if os.path.isfile(excel_file) else True
        if not must_add_headers:
            wb =load_workbook(excel_file)
            ws = wb.active

        else:
            wb = openpyxl.Workbook() 
        # Sheet_name = wb.sheetnames
            ws = wb.active
            ws.append(
                ('Company Site','Skill','Job Title', 'Job Link', 'Source')
            )
        # ws.append(
        #     ('Company Site','Skill','Job Title', 'Job Link')
        # )
        for i in data: 
            j =(
                i['Company Site'],i['Skill'],i['Job Title'],i['Job Link'],i['Source'] 
            )  
            ws.append(j)
        # wb.save(f'{excel_file}.xlsx')
        wb.save(excel_file)
        
    except Exception as e:
        pass
# data = [ 
#  {"Skill": "skill 123","Job Title" : "java python Ruby", "Job Link": "www.askjndjas"},
#  {"Skill": "skill 123","Job Title" : "java python Ruby", "Job Link": "www.askjndjas"},
#  {"Skill": "skill 456","Job Title" : "java python Ruby", "Job Link": "www.askjndjas"}
 
# ]


def fetchSkills():
    try:
        result = requests.get("http://crawler.allshoreresources.com:8088/api/skills")
        result = result.json()
        return result
    except Exception as e:
        # print(e)
        pass


def saveToCsvFile(data, result):
    field_names = ['Skill', 'Job Title', 'Job Link']
    # print(str(company_name).split('.')[0])
    try:  
        csv_file = f"{result.split('.')[0]}.csv"
        with open(csv_file, 'w', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=field_names)
            # Write the header row
            writer.writeheader()
            # Write the data rows
            writer.writerows(data)
        print('Data has been successfully written to', csv_file)
    except Exception as e:
        # print(e)
        pass  

def saveTAllDataCsv(data, result):
    field_names = ['Company Site','Skill', 'Job Title', 'Job Link']
    csv_file = "CompleteData.csv"
    try:  
        must_add_headers = False if os.path.isfile(csv_file) else True
        with open(csv_file, 'a', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=field_names)
            if must_add_headers:
                writer.writeheader()
            writer.writerows(data)
                
        print('Data has been successfully written to', csv_file)
    except Exception as e:
        # print(e)
        pass  


def getElementByTag(e, tagName):
    
    if e == None:
        return None
    if e.tag_name == "body":
        return None
    elif e.tag_name == tagName:
        # print(e)
        # print(e.tag_name)
        # print(e.get_attribute('href'))
        # print('++++++++++++++++++++++++++')
        
        return e
    else:
        try:
            
            e = e.find_element(By.XPATH, '..')
            # print('---------------------------')
            # print(e)
            # print('----------------------')
            # print('looking parent element')
            # print(e.tag_name)
            # print(e.tag_name)
            return getElementByTag(e, tagName)
        except Exception as e:
            # print('GET ELEMENT BY TAG')
            # print(e)
            return None
    
# def getJobsfromCurrentPage(driver, data):
#     for skill in SKILLS:
#         skill_xpath = f"//*[contains(text(), '{skill}')]"
#         print(skill_xpath)
#         try:
#             element = driver.find_element(By.XPATH, skill_xpath) 
        
#             if element: 
#                 print('skill is there')
#                 print('skill is there')
#                 print(element.tag_name)
#                 element_link = getElementByTag(element, 'a')
#                 element_link.get_attribute('href')
                
#                 print(element.get_attribute("innerHTML"))
#                 d = {"Skill": skill,"Job Title" : element.get_attribute("innerHTML"), "Job Link": element_link.get_attribute('href')}
#                 data.append(d)
#         except Exception as e:
#             print(e)
            
#     # driver.find_element(By.XPATH, f"//a[contains(text(), 'next')]")
        
        
#     return data
# getJobsfromCurrentPage

                            
#                             current_jobs_data = getJobsfromCurrentPage(driver, data)
#                             data.append(current_jobs_data)

































    # if request.method == 'POST':
    #     try:
    #         company_name = request.FILES['company']
    #     except MultiValueDictKeyError:
    #         return render(request, 'demo/form.html')
    #     splitting = str(company_name).split('.')[-1]
    #     if (splitting == 'xls'):
    #         data = xls_get(company_name, column=2)
    #     elif (splitting == 'xlsx'):
    #         data = xlsx_get(company_name, column=2)
    #     else:
    #         return render(request, 'demo/form.html')
    #     sheets = data.keys()
    #     sheet = 0
    #     for i in data.keys():
    #         sheet = i
    #         break
    #     companies = data[sheet]
    #     # print(companies)
    #     job_titles = []
    #     websites = []
    #     websites_not_found = []
        
    #     iteration = 0
        # for company in companies[1:]:
        #     if company[1] == "Direct Employer" :
        #         # and company[0] not in job_titles
        #         x = re.sub("\" | Inc.| Llc| LLC|\'|, LLC|, Inc.|, Inc.|, Llc| Inc|, Inc| Inc.| Inc| INC|, LLC.|, Ltd|\.|,","", company[0])
        #         if x not in job_titles: 
        #             job_titles.append(x)
        #             url = "https://autocomplete.clearbit.com/v1/companies/suggest?query="+x
        #             result = requests.get(url)
        #             result = result.json()
        #             if result:
        #                 websites.append(result[0]['domain'])
        #             else:
        #                 websites_not_found.append(x)